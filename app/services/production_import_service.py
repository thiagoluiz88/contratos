from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models import Contract, Operator, ProductionImportBatch, ProductionImportLayout, ProductionRecord


BATCH_PENDING = "pendente"
BATCH_PROCESSED = "processado"
BATCH_ERROR = "erro"
BATCH_CANCELLED = "cancelado"
VALIDATION_VALID = "valido"
VALIDATION_PENDING = "pendente"
MAX_IMPORT_SIZE = 20 * 1024 * 1024

STANDARD_COLUMNS = {
    "operadora", "contrato", "competencia", "data_atendimento", "categoria", "item", "descricao",
    "quantidade", "unidade", "valor_faturado", "valor_pago", "valor_glosado", "custo",
    "guia", "conta", "atendimento", "paciente_referencia",
}
REQUIRED_COLUMNS = {"operadora", "contrato", "competencia", "data_atendimento", "categoria", "item", "descricao", "quantidade", "unidade", "valor_faturado", "valor_pago", "valor_glosado", "custo"}
COLUMN_ALIASES = {"cnpj_operadora": "operadora", "nome_operadora": "operadora", "numero_contrato": "contrato", "mes_competencia": "competencia", "data_servico": "data_atendimento", "valor_glosa": "valor_glosado", "valor_custo": "custo", "paciente": "paciente_referencia"}


@dataclass(slots=True)
class ProductionAuditEvent:
    action: str
    entity_type: str | None = None
    entity_id: int | None = None
    details: str | None = None
    success: bool = True


def fold_column(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    folded = "".join(char for char in normalized if not unicodedata.combining(char)).lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", folded).strip("_")


def normalize_columns(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {}
    for key, value in row.items():
        column = COLUMN_ALIASES.get(fold_column(key), fold_column(key))
        normalized[column] = value.strip() if isinstance(value, str) else value
    return normalized


def validate_import_file(path: Path) -> None:
    if path.suffix.lower() != ".csv":
        raise ValueError("Nesta versão, envie um arquivo CSV.")
    if not path.exists() or path.stat().st_size == 0:
        raise ValueError("Arquivo CSV vazio ou não encontrado.")
    if path.stat().st_size > MAX_IMPORT_SIZE:
        raise ValueError("Arquivo CSV excede o limite de 20 MB.")
    if b"\x00" in path.read_bytes()[:4096]:
        raise ValueError("Arquivo CSV inválido ou binário.")


def _read_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Não foi possível identificar a codificação do CSV.")


def parse_csv(path: Path, layout=None) -> list[dict[str, Any]]:
    validate_import_file(path)
    if layout and layout.encoding:
        try: text = path.read_text(encoding=layout.encoding)
        except (LookupError, UnicodeDecodeError) as exc: raise ValueError("Codificação configurada no layout não pôde ler o CSV.") from exc
    else:
        text = _read_text(path)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    if layout and layout.delimiter:
        reader = csv.DictReader(text.splitlines(), delimiter=layout.delimiter)
    else:
        reader = csv.DictReader(text.splitlines(), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("CSV sem cabeçalho.")
    raw_rows = [dict(row) for row in reader]
    if layout:
        from app.services.production_layout_service import apply_layout_mapping, detect_missing_required_columns
        missing = detect_missing_required_columns(list(reader.fieldnames), layout)
        if missing: raise ValueError(f"Arquivo sem colunas obrigatórias do layout: {', '.join(sorted(missing))}.")
        return [apply_layout_mapping(row, layout) for row in raw_rows]
    headers = {COLUMN_ALIASES.get(fold_column(name), fold_column(name)) for name in reader.fieldnames if name}
    missing = REQUIRED_COLUMNS - headers
    if missing: raise ValueError(f"CSV sem colunas obrigatórias: {', '.join(sorted(missing))}.")
    return [normalize_columns(row) for row in raw_rows]


def parse_excel(path: Path, layout=None) -> list[dict[str, Any]]:
    if path.suffix.lower() != ".xlsx": raise ValueError("Formato Excel legado .xls não é suportado; envie .xlsx.")
    if not path.exists() or path.stat().st_size == 0 or path.stat().st_size > MAX_IMPORT_SIZE: raise ValueError("Arquivo Excel vazio, ausente ou acima de 20 MB.")
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator)]
        raw_rows = [{headers[index]: (value.isoformat() if hasattr(value, "isoformat") else value) for index, value in enumerate(values)} for values in iterator]
        workbook.close()
    except (StopIteration, OSError, ValueError) as exc: raise ValueError("Não foi possível ler a primeira aba do arquivo Excel.") from exc
    if layout:
        from app.services.production_layout_service import apply_layout_mapping, detect_missing_required_columns
        missing = detect_missing_required_columns(headers, layout)
        if missing: raise ValueError(f"Excel sem colunas obrigatórias do layout: {', '.join(sorted(missing))}.")
        return [apply_layout_mapping(row, layout) for row in raw_rows]
    normalized_headers = {COLUMN_ALIASES.get(fold_column(name), fold_column(name)) for name in headers}
    missing = REQUIRED_COLUMNS - normalized_headers
    if missing: raise ValueError(f"Excel sem colunas obrigatórias: {', '.join(sorted(missing))}.")
    return [normalize_columns(row) for row in raw_rows]


def parse_decimal(value: Any) -> Decimal | None:
    prepared = str(value or "").strip().replace("R$", "").replace(" ", "")
    if not prepared:
        return None
    if "," in prepared:
        prepared = prepared.replace(".", "").replace(",", ".")
    try:
        return Decimal(prepared)
    except InvalidOperation:
        return None


def parse_date(value: Any) -> date | None:
    prepared = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m", "%m/%Y"):
        try:
            parsed = datetime.strptime(prepared, pattern).date()
            return parsed.replace(day=1) if pattern in ("%Y-%m", "%m/%Y") else parsed
        except ValueError:
            continue
    return None


def competence_date(value: Any) -> date | None:
    parsed = parse_date(value)
    return parsed.replace(day=1) if parsed else None


def resolve_operator(db: Session, value: Any) -> Operator | None:
    prepared = str(value or "").strip()
    if not prepared:
        return None
    digits = re.sub(r"\D", "", prepared)
    operator = None
    if len(digits) >= 8:
        operator = db.query(Operator).filter(func.regexp_replace(Operator.tax_id, "[^0-9]", "", "g") == digits).first()
    return operator or db.query(Operator).filter(func.lower(Operator.name) == prepared.lower()).first()


def resolve_contract(db: Session, value: Any, operator: Operator | None = None) -> Contract | None:
    prepared = str(value or "").strip()
    if not prepared:
        return None
    query = db.query(Contract).filter(Contract.status == "active")
    if operator:
        query = query.filter(Contract.operator_id == operator.id)
    if prepared.isdigit():
        by_id = query.filter(Contract.id == int(prepared)).first()
        if by_id:
            return by_id
    return query.filter(or_(func.lower(Contract.contract_number) == prepared.lower(), func.lower(Contract.contract_name) == prepared.lower())).first()


def validate_row(row: dict[str, Any], operator: Operator | None, contract: Contract | None) -> tuple[str, list[str]]:
    messages = []
    if not competence_date(row.get("competencia")):
        messages.append("Competência inválida ou ausente.")
    if not parse_date(row.get("data_atendimento")):
        messages.append("Data de atendimento inválida ou ausente.")
    if not str(row.get("categoria") or "").strip():
        messages.append("Categoria ausente.")
    if not str(row.get("item") or "").strip():
        messages.append("Item ausente.")
    quantity = parse_decimal(row.get("quantidade"))
    if quantity is None or quantity <= 0:
        messages.append("Quantidade inválida ou não positiva.")
    if not operator:
        messages.append("Operadora não localizada.")
    if not contract:
        messages.append("Contrato não localizado.")
    financial_fields = ("valor_faturado", "valor_pago", "valor_glosado", "custo")
    for field in financial_fields:
        raw = str(row.get(field) or "").strip()
        if raw and parse_decimal(raw) is None:
            messages.append(f"{field} inválido.")
    return (VALIDATION_VALID if not messages else VALIDATION_PENDING), messages


def patient_hash(value: Any) -> str | None:
    prepared = str(value or "").strip()
    return hashlib.sha256(prepared.encode("utf-8")).hexdigest() if prepared else None


def create_import_batch(*, batch_name: str, source_type: str = "csv", source_system: str = "planilha", original_filename: str | None = None, file_path: str | None = None, imported_by: str | None = None, notes: str | None = None, layout_id: int | None = None) -> tuple[ProductionImportBatch, list[ProductionAuditEvent]]:
    db = SessionLocal()
    try:
        batch = ProductionImportBatch(batch_name=batch_name.strip(), source_type=source_type, source_system=source_system, original_filename=original_filename, file_path=file_path, imported_by=imported_by, notes=notes, import_status=BATCH_PENDING, layout_id=layout_id)
        if not batch.batch_name:
            raise ValueError("Nome do lote é obrigatório.")
        db.add(batch)
        db.commit()
        db.refresh(batch)
        return batch, [ProductionAuditEvent("production_import_batch_created", "production_import_batch", batch.id, batch.batch_name)]
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def create_production_records(batch_id: int, rows: list[dict[str, Any]]) -> tuple[ProductionImportBatch, list[ProductionAuditEvent]]:
    db = SessionLocal()
    events: list[ProductionAuditEvent] = []
    try:
        batch = db.query(ProductionImportBatch).filter(ProductionImportBatch.id == batch_id).first()
        if not batch:
            raise ValueError("Lote não encontrado.")
        if batch.import_status != BATCH_PENDING or batch.records:
            raise ValueError("Lote já processado ou indisponível para reprocessamento.")
        valid = invalid = 0
        for row_number, row in enumerate(rows, start=2):
            operator = resolve_operator(db, row.get("operadora"))
            contract = resolve_contract(db, row.get("contrato"), operator)
            status, messages = validate_row(row, operator, contract)
            record = ProductionRecord(
                batch_id=batch.id, operator_id=operator.id if operator else None, contract_id=contract.id if contract else None,
                patient_identifier_hash=patient_hash(row.get("paciente_referencia")), attendance_reference=str(row.get("atendimento") or "").strip() or None,
                account_reference=str(row.get("conta") or "").strip() or None, guide_reference=str(row.get("guia") or "").strip() or None,
                service_date=parse_date(row.get("data_atendimento")), competence_month=competence_date(row.get("competencia")),
                category=str(row.get("categoria") or "").strip() or None, item=str(row.get("item") or "").strip() or None,
                description=str(row.get("descricao") or "").strip() or None, quantity=parse_decimal(row.get("quantidade")), unit=str(row.get("unidade") or "").strip() or None,
                billed_value=parse_decimal(row.get("valor_faturado")), paid_value=parse_decimal(row.get("valor_pago")), denied_value=parse_decimal(row.get("valor_glosado")), cost_value=parse_decimal(row.get("custo")),
                source_row_number=row_number, validation_status=status, validation_message=" ".join(messages) or None,
            )
            db.add(record)
            db.flush()
            if status == VALIDATION_VALID:
                valid += 1
            else:
                invalid += 1
                events.append(ProductionAuditEvent("production_record_invalid_detected", "production_record", record.id, f"Lote #{batch.id}; linha {row_number}.", False))
        batch.total_rows = len(rows)
        batch.valid_rows = valid
        batch.invalid_rows = invalid
        batch.import_status = BATCH_PROCESSED
        batch.processed_at = datetime.utcnow()
        batch.error_message = None
        events.append(ProductionAuditEvent("production_import_batch_processed", "production_import_batch", batch.id, f"{len(rows)} linha(s); {valid} válida(s); {invalid} pendente(s)."))
        db.commit()
        db.refresh(batch)
        return batch, events
    except Exception:
        db.rollback()
        batch = db.query(ProductionImportBatch).filter(ProductionImportBatch.id == batch_id).first()
        if batch and batch.import_status == BATCH_PENDING:
            batch.import_status = BATCH_ERROR
            batch.error_message = "Falha ao processar o arquivo de produção."
            db.commit()
        raise
    finally:
        db.close()


def import_file_to_batch(batch_id: int, path: Path) -> tuple[ProductionImportBatch, list[ProductionAuditEvent]]:
    db = SessionLocal()
    try:
        batch = db.query(ProductionImportBatch).filter(ProductionImportBatch.id == batch_id).first()
        if not batch: raise ValueError("Lote não encontrado.")
        layout = batch.layout
        raw_rows = parse_excel(path, layout) if path.suffix.lower() == ".xlsx" else parse_csv(path, layout)
        if layout:
            # Releitura mínima apenas para metadados de cabeçalho, sem persistir conteúdo sensível.
            if path.suffix.lower() == ".xlsx":
                from openpyxl import load_workbook
                book = load_workbook(path, read_only=True, data_only=True); columns = [str(v or "") for v in next(book.worksheets[0].iter_rows(values_only=True))]; book.close()
            else:
                first = _read_text(path).splitlines()[0]; columns = next(csv.reader([first], delimiter=layout.delimiter or ";"))
            from app.services.production_layout_service import build_layout_preview
            preview = build_layout_preview([{column: None for column in columns}], layout, 0)
            summary = {"layout_id": layout.id, "layout_name": layout.name, "recognized_columns": preview["recognized_columns"], "ignored_columns": preview["unmapped_columns"], "missing_required_fields": preview["missing_required_fields"]}
        else:
            summary = {"layout_id": None, "layout_name": "Aliases padrão", "recognized_columns": sorted(REQUIRED_COLUMNS), "ignored_columns": [], "missing_required_fields": []}
        batch.import_summary_json = summary; db.commit()
    finally: db.close()
    rows = raw_rows
    batch, events = create_production_records(batch_id, rows)
    action = "production_excel_file_imported" if path.suffix.lower() == ".xlsx" else "production_import_file_received"
    layout_event = [ProductionAuditEvent("production_import_with_layout", "production_import_batch", batch.id, f"Layout #{batch.layout_id}.")] if batch.layout_id else []
    return batch, [ProductionAuditEvent(action, "production_import_batch", batch.id, batch.original_filename), *layout_event, *events]


def import_csv_to_batch(batch_id: int, path: Path) -> tuple[ProductionImportBatch, list[ProductionAuditEvent]]:
    return import_file_to_batch(batch_id, path)


def build_import_preview(path: Path, *, layout_id: int | None = None, delimiter: str | None = None, encoding: str | None = None, sheet_name: str | None = None, limit: int = 50) -> dict[str, Any]:
    db=SessionLocal()
    try:
        layout=db.query(ProductionImportLayout).filter_by(id=layout_id).first() if layout_id else None
        extension=path.suffix.lower(); sheets=[]; used_sheet=None
        if extension==".xlsx":
            from openpyxl import load_workbook
            if not path.exists() or path.stat().st_size>MAX_IMPORT_SIZE: raise ValueError("Arquivo Excel inválido ou acima de 20 MB.")
            book=load_workbook(path,read_only=True,data_only=True);sheets=book.sheetnames
            sheet=book[sheet_name] if sheet_name in sheets else book.worksheets[0];used_sheet=sheet.title
            iterator=sheet.iter_rows(values_only=True);columns=[str(v or "").strip() for v in next(iterator)]
            raw=[{columns[i]:(v.isoformat() if hasattr(v,"isoformat") else v) for i,v in enumerate(values)} for values in list(iterator)[:limit]];book.close()
        elif extension==".csv":
            validate_import_file(path)
            text=path.read_text(encoding=encoding) if encoding else _read_text(path)
            lines=text.splitlines(); chosen=delimiter
            if not chosen:
                try: chosen=csv.Sniffer().sniff(text[:4096],delimiters=";,\t").delimiter
                except csv.Error: chosen=";"
            reader=csv.DictReader(lines,delimiter=chosen);columns=list(reader.fieldnames or []);raw=[dict(row) for _,row in zip(range(limit),reader)]
        else: raise ValueError("Preview aceita somente CSV ou Excel .xlsx.")
        if not columns: raise ValueError("Arquivo sem cabeçalho detectável.")
        if layout:
            from app.services.production_layout_service import apply_layout_mapping,detect_missing_required_columns,detect_unmapped_columns
            normalized=[apply_layout_mapping(row,layout) for row in raw];ignored=detect_unmapped_columns(columns,layout);missing=detect_missing_required_columns(columns,layout);recognized=[c for c in columns if c not in ignored]
        else:
            normalized=[normalize_columns(row) for row in raw];header_targets={COLUMN_ALIASES.get(fold_column(c),fold_column(c)) for c in columns};missing=sorted(REQUIRED_COLUMNS-header_targets);recognized=[c for c in columns if COLUMN_ALIASES.get(fold_column(c),fold_column(c)) in STANDARD_COLUMNS];ignored=[c for c in columns if c not in recognized]
        previews=[];valid=invalid=0
        for index,row in enumerate(normalized,start=2):
            operator=resolve_operator(db,row.get("operadora"));contract=resolve_contract(db,row.get("contrato"),operator);status,messages=validate_row(row,operator,contract)
            safe=dict(row)
            if safe.get("paciente_referencia"): safe["paciente_referencia_hash"]=patient_hash(safe.pop("paciente_referencia"))
            previews.append({"source_row_number":index,"status":status,"messages":messages,"data":safe,"operator":operator.name if operator else None,"contract":contract.contract_name if contract else None})
            valid+=status==VALIDATION_VALID;invalid+=status!=VALIDATION_VALID
        return {"columns":columns,"recognized_columns":recognized,"ignored_columns":ignored,"missing_required_fields":missing,"rows":previews,"valid_rows":valid,"invalid_rows":invalid,"layout":layout,"compatible":not missing,"sheet_names":sheets,"used_sheet":used_sheet,"analyzed_rows":len(previews)}
    finally: db.close()


def build_import_summary(db: Session, batch_id: int) -> dict[str, Any]:
    records = db.query(ProductionRecord).filter(ProductionRecord.batch_id == batch_id).all()
    def total(field: str):
        return sum((Decimal(str(getattr(record, field))) for record in records if getattr(record, field) is not None), Decimal("0"))
    return {"total_rows": len(records), "valid_rows": sum(record.validation_status == VALIDATION_VALID for record in records), "invalid_rows": sum(record.validation_status != VALIDATION_VALID for record in records), "billed_value": total("billed_value"), "paid_value": total("paid_value"), "denied_value": total("denied_value"), "cost_value": total("cost_value"), "records_with_cost": sum(record.cost_value is not None for record in records)}


def cancel_import_batch(batch_id: int, *, cancelled_by: str | None = None) -> tuple[ProductionImportBatch, list[ProductionAuditEvent]]:
    db = SessionLocal()
    try:
        batch = db.query(ProductionImportBatch).filter(ProductionImportBatch.id == batch_id).first()
        if not batch:
            raise ValueError("Lote não encontrado.")
        if batch.import_status not in {BATCH_PENDING, BATCH_ERROR} or batch.records:
            raise ValueError("Somente lote pendente ou com erro, sem registros consolidados, pode ser cancelado.")
        batch.import_status = BATCH_CANCELLED
        batch.notes = " | ".join(part for part in (batch.notes, f"Cancelado por {cancelled_by or 'usuário não informado'} em {datetime.utcnow().isoformat(timespec='seconds')}.") if part)
        db.commit()
        db.refresh(batch)
        return batch, [ProductionAuditEvent("production_import_batch_cancelled", "production_import_batch", batch.id, batch.batch_name)]
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
